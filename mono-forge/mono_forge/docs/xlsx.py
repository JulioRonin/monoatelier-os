"""cutlist.xlsx y herrajes.xlsx — los documentos que baja el taller.

Todo sale del project.json. El cubrecanto va pieza por pieza y lado por lado
(nunca "todos los cantos" en bloque: infla ~40% el costo real).
"""

from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..constants import HOJA_ANCHO, HOJA_LARGO, HOJA_M2, KERF
from ..costing import catalogo_herrajes, catalogo_materiales
from ..cutlist import nesting, resumen
from ..models import Project

PRIMARY = "FF427A6E"
GRIS = "FFF3F4F6"
_TITULO = Font(bold=True, color="FFFFFFFF", size=10)
_FILL = PatternFill("solid", fgColor=PRIMARY)
_BORDE = Border(bottom=Side(style="thin", color="FFE5E7EB"))


def _encabezar(ws, columnas: list[tuple[str, int]], fila: int = 1) -> None:
    for i, (titulo, ancho) in enumerate(columnas, start=1):
        c = ws.cell(row=fila, column=i, value=titulo)
        c.font = _TITULO
        c.fill = _FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = ancho
    ws.freeze_panes = ws.cell(row=fila + 1, column=1)


def _cantos_txt(p) -> str:
    lados = [k for k, v in p.cantos.items() if v]
    return ", ".join(lados) if lados else "—"


def cutlist_xlsx(project: Project, destino: str) -> str:
    """Piezas + cubrecanto + perforaciones + nesting + resumen."""
    wb = Workbook()

    # ── Piezas ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Piezas"
    _encabezar(ws, [("Módulo", 12), ("Pieza", 30), ("Cant", 6), ("Largo mm", 10),
                    ("Ancho mm", 10), ("Esp mm", 8), ("Material", 18),
                    ("Veta", 10), ("Cantos", 22), ("Rol estructural", 18),
                    ("Área m²", 9)])
    fila = 2
    contenedores = [(m.id, m.panels) for m in project.modules] + \
                   [(t.id, t.panels) for t in project.tramos]
    for mid, panels in contenedores:
        for p in panels:
            if p.accesorio:
                continue    # se compra, no se corta
            for col, val in enumerate([
                mid, p.name, int(p.cantidad), p.largo, p.ancho, p.espesor,
                p.material, p.veta, _cantos_txt(p), p.rol_estructural,
                round(p.area_m2, 4)], start=1):
                cell = ws.cell(row=fila, column=col, value=val)
                cell.border = _BORDE
            fila += 1
    ws.cell(row=fila + 1, column=1, value="ÁREA TOTAL m²").font = Font(bold=True)
    ws.cell(row=fila + 1, column=11,
            value=round(sum(p.area_m2 for p in project.piezas_de_corte()), 3)
            ).font = Font(bold=True)

    # ── Cubrecanto ──────────────────────────────────────────────────
    ws = wb.create_sheet("Cubrecanto")
    _encabezar(ws, [("Pieza", 30), ("Cant", 6), ("Lados encintados", 24),
                    ("ml por pieza", 12), ("ml total", 12), ("Material", 18),
                    ("Aplicación", 14)])
    fila = 2
    mats = catalogo_materiales()
    for p in project.piezas_de_corte():
        if p.ml_canto <= 0:
            continue
        desc = (mats.get(p.material, {}).get("descripcion", "") + p.material).lower()
        aplicacion = "MANUAL" if "brillo" in desc else "máquina"
        for col, val in enumerate([
            p.name, int(p.cantidad), _cantos_txt(p),
            round(p.ml_canto / max(p.cantidad, 1), 3), round(p.ml_canto, 3),
            p.material, aplicacion], start=1):
            ws.cell(row=fila, column=col, value=val).border = _BORDE
        fila += 1
    ws.cell(row=fila + 1, column=1, value="TOTAL ml").font = Font(bold=True)
    ws.cell(row=fila + 1, column=5,
            value=round(sum(p.ml_canto for p in project.piezas_de_corte()), 2)
            ).font = Font(bold=True)
    ws.cell(row=fila + 3, column=1,
            value="El alto brillo SIEMPRE lleva cintilla PVC pegada a mano.")

    # ── Perforaciones ───────────────────────────────────────────────
    ws = wb.create_sheet("Perforaciones")
    _encabezar(ws, [("Pieza", 30), ("Tipo", 14), ("Ø mm", 8), ("X mm", 10),
                    ("Y mm", 10), ("Prof mm", 10), ("Cara", 12), ("Nota", 32)])
    fila = 2
    for p in project.piezas_de_corte():
        for d in p.drilling:
            for col, val in enumerate([p.name, d.tipo, d.diametro, d.x, d.y,
                                       d.profundidad, d.cara, d.nota], start=1):
                ws.cell(row=fila, column=col, value=val).border = _BORDE
            fila += 1
    if fila == 2:
        ws.cell(row=2, column=1, value="Sin perforaciones registradas.")

    # ── Nesting ─────────────────────────────────────────────────────
    ws = wb.create_sheet("Nesting")
    _encabezar(ws, [("Material", 20), ("Hoja #", 8), ("Pieza", 30), ("X mm", 10),
                    ("Y mm", 10), ("Largo mm", 10), ("Ancho mm", 10)])
    fila = 2
    hojas = nesting(project.piezas_de_corte())
    for material, hs in hojas.items():
        for h in hs:
            for c in h.piezas:
                for col, val in enumerate([material, h.indice + 1, c.panel,
                                           round(c.x, 1), round(c.y, 1),
                                           c.largo, c.ancho], start=1):
                    ws.cell(row=fila, column=col, value=val).border = _BORDE
                fila += 1

    # ── Resumen ─────────────────────────────────────────────────────
    ws = wb.create_sheet("Resumen")
    r = resumen(project)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 22
    filas = [
        ("Proyecto", project.nombre),
        ("Cliente", project.cliente),
        ("Hoja estándar", f"{HOJA_LARGO} × {HOJA_ANCHO} mm ({HOJA_M2:.4f} m²)"),
        ("Kerf (ancho de sierra)", f"{KERF} mm"),
        ("Área total de paneles", f"{r['area_paneles_m2']} m²"),
        ("Cubrecanto total", f"{r['ml_cubrecanto']} ml"),
        ("", ""),
    ]
    for material, n in r["hojas_por_material"].items():
        filas.append((f"Hojas de {material}",
                      f"{n} — aprovechamiento {r['aprovechamiento'][material]*100:.1f}%"))
    if r["alertas"]:
        filas.append(("", ""))
        filas.append(("ALERTAS DE APROVECHAMIENTO", ""))
        filas += [("", a) for a in r["alertas"]]
    for i, (k, v) in enumerate(filas, start=1):
        ws.cell(row=i, column=1, value=k).font = Font(bold=bool(k))
        ws.cell(row=i, column=2, value=v)

    ruta = os.path.join(destino, "cutlist.xlsx")
    wb.save(ruta)
    return ruta


def herrajes_xlsx(project: Project, destino: str) -> str:
    """Lista consolidada de compra, con precios del catálogo cuando existen."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Herrajes"
    _encabezar(ws, [("SKU", 18), ("Descripción", 46), ("Cantidad", 10),
                    ("Unidad", 8), ("Costo unit.", 12), ("Importe", 12),
                    ("Proveedor", 18)])

    cat = catalogo_herrajes()
    fila = 2
    total = 0.0
    sin_precio: list[str] = []
    for h in project.hardware_consolidado().values():
        info = cat.get(h.sku, {})
        try:
            costo = float(info.get("costo_unit") or 0)
        except ValueError:
            costo = 0.0
        importe = costo * h.cantidad if costo else None
        if importe is None:
            sin_precio.append(h.sku)
        else:
            total += importe
        for col, val in enumerate([h.sku, h.descripcion, h.cantidad, h.unidad,
                                   costo or None, importe,
                                   info.get("proveedor", "")], start=1):
            ws.cell(row=fila, column=col, value=val).border = _BORDE
        fila += 1

    ws.cell(row=fila + 1, column=2, value="TOTAL (sólo SKUs con precio)").font = Font(bold=True)
    ws.cell(row=fila + 1, column=6, value=round(total, 2)).font = Font(bold=True)
    if sin_precio:
        ws.cell(row=fila + 3, column=1, value="SKUs SIN PRECIO EN CATÁLOGO:"
                ).font = Font(bold=True, color="FF991B1B")
        ws.cell(row=fila + 3, column=2, value=", ".join(sin_precio))
        ws.cell(row=fila + 4, column=2,
                value="Completa catalog/herrajes.csv para que el costeo cierre.")

    # bisagras: el tipo NO es intercambiable al comprar
    ws2 = wb.create_sheet("Nota de compra")
    ws2.column_dimensions["A"].width = 100
    notas = [
        "El tipo de bisagra (recta / semicurva / curva) lo determina la modulación,",
        "no es intercambiable al comprar: recta = sobreposición total, semicurva =",
        "media sobreposición (dos puertas comparten divisor), curva = puerta interior.",
        "",
        "Corredera estándar del taller: 500mm. No subir a 550 aunque el módulo dé.",
    ]
    for i, n in enumerate(notas, start=1):
        ws2.cell(row=i, column=1, value=n)

    ruta = os.path.join(destino, "herrajes.xlsx")
    wb.save(ruta)
    return ruta
