"""Red de seguridad de los entregables.

Los documentos son la cara del sistema ante el cliente y el taller: si un
número no coincide con el project.json, el error se propaga a la obra.
"""

import openpyxl
import pytest

from mono_forge.costing import Tarifas, comparar_proveedores, costear
from mono_forge.cutlist import resumen
from mono_forge.docs import ESPERADOS, generar_todo, verificar
from mono_forge.docs.manual import pasos_de
from mono_forge.generators.base import gabinete_base
from mono_forge.generators.cubierta import cubierta
from mono_forge.generators.superior import alacena
from mono_forge.models import Project, Tramo
from mono_forge.rules.posicion import colocar


@pytest.fixture
def proyecto():
    p = Project(cliente="TEST", nombre="entregables")
    p.modules.append(gabinete_base("B01", ancho=600))
    p.modules.append(gabinete_base("B02", ancho=1240, puertas=2, tarja=True,
                                   entrepanos=0))
    p.modules.append(alacena("A01", ancho=600))
    t = Tramo(id="T1", muro="A", modulos=["B01", "B02"])
    panels, notas = cubierta("T1", 1840)
    t.panels += panels
    t.notas += notas
    p.tramos.append(t)
    colocar(p)
    return p


@pytest.fixture
def entregables(proyecto, tmp_path):
    destino = str(tmp_path / "deliverables")
    rutas = generar_todo(proyecto, destino, Tarifas(
        canto_maquina_ml=12, canto_manual_ml=45, mano_obra_modulo=850, margen=0.35))
    return proyecto, destino, rutas


def test_se_generan_los_seis_documentos(entregables):
    _, _, rutas = entregables
    assert set(rutas) == {"cutlist.xlsx", "herrajes.xlsx", "cotizacion.pdf",
                          "manual_ensamble.pdf", "entrega.pdf", "costos_internos.pdf"}
    for ruta in rutas.values():
        with open(ruta, "rb") as f:
            assert len(f.read()) > 1000, ruta


def test_verificacion_area_cuadra(entregables):
    proyecto, destino, _ = entregables
    v = verificar(proyecto, destino)
    # el área del cutlist debe reproducir la del JSON (la tolerancia es 1%)
    assert v["area_json_m2"] == pytest.approx(v["area_cutlist_m2"], rel=0.01)
    assert v["problemas"] == []
    # sólo pueden faltar los que produce Blender, no los documentos
    assert set(v["faltantes"]) <= {"modelo.blend", "preview.glb"}
    assert len(ESPERADOS) == 8


def test_cutlist_xlsx_refleja_el_json(entregables):
    proyecto, _, rutas = entregables
    wb = openpyxl.load_workbook(rutas["cutlist.xlsx"])
    assert {"Piezas", "Cubrecanto", "Perforaciones", "Nesting", "Resumen"} \
        <= set(wb.sheetnames)

    ws = wb["Piezas"]
    filas = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[1]]
    esperadas = {p.name for p in proyecto.all_panels()}
    assert {f[1] for f in filas} == esperadas

    # el área total impresa debe ser la del modelo de datos
    # (la hoja redondea a 4 decimales por pieza: esa es la única diferencia tolerable)
    area_json = sum(p.area_m2 for p in proyecto.all_panels())
    area_hoja = sum(f[10] for f in filas)
    assert area_hoja == pytest.approx(area_json, rel=1e-3)


def test_cubrecanto_no_se_aplica_en_bloque(entregables):
    """Aplicar los 4 cantos a todo infla ~40%: la hoja debe listar sólo los reales."""
    proyecto, _, rutas = entregables
    wb = openpyxl.load_workbook(rutas["cutlist.xlsx"])
    ws = wb["Cubrecanto"]
    # sólo las filas de pieza (la de TOTAL no trae cantidad numérica)
    total = sum(r[4] for r in ws.iter_rows(min_row=2, values_only=True)
                if isinstance(r[1], int) and isinstance(r[4], (int, float)))
    assert total == pytest.approx(resumen(proyecto)["ml_cubrecanto"], rel=1e-3)

    # ninguna pieza debe llevar los 4 cantos "por si acaso": sólo los frentes
    piezas_4_cantos = [r[0] for r in ws.iter_rows(min_row=2, values_only=True)
                       if isinstance(r[1], int) and r[2] and r[2].count(",") == 3]
    assert all("puerta" in p or "frente" in p for p in piezas_4_cantos)


def test_herrajes_xlsx_consolida_por_sku(entregables):
    proyecto, _, rutas = entregables
    wb = openpyxl.load_workbook(rutas["herrajes.xlsx"])
    ws = wb["Herrajes"]
    filas = {r[0]: r[2] for r in ws.iter_rows(min_row=2, values_only=True) if r[0]}
    for sku, item in proyecto.hardware_consolidado().items():
        assert filas[sku] == pytest.approx(item.cantidad)


def test_pasos_de_ensamble_explican_el_porque(proyecto):
    base = proyecto.modules[0]
    pasos = pasos_de(base)
    titulos = " ".join(t for t, _, _ in pasos).lower()
    assert "apoyar los laterales" in titulos      # la regla estructural
    assert all(porque.strip() for _, _, porque in pasos)   # ningún paso sin razón
    # el mueble colgado NO debe decir "apoyar sobre la base"
    colgado = " ".join(t for t, _, _ in pasos_de(proyecto.modules[2])).lower()
    assert "capturar techo y piso" in colgado
    assert "apoyar los laterales" not in colgado


def test_el_margen_nunca_sale_en_documentos_de_cliente(entregables):
    """El margen sólo vive en costos_internos.pdf."""
    proyecto, _, rutas = entregables
    costos = costear(proyecto, Tarifas(margen=0.35))
    marcadores = [b"margen", b"Margen", b"MARGEN",
                  f"{costos['costo_directo']:,.2f}".encode()]
    for doc in ("cotizacion.pdf", "entrega.pdf"):
        with open(rutas[doc], "rb") as f:
            crudo = f.read()
        for m in marcadores:
            assert m not in crudo, f"{doc} filtra información interna: {m!r}"


def test_simulacion_por_proveedor_diferencia_precios(proyecto):
    """Si todos los proveedores dan el mismo número, la comparación es inútil."""
    comparacion = comparar_proveedores(proyecto, Tarifas(margen=0.35))
    assert len(comparacion) >= 2
    assert len(set(comparacion.values())) > 1
    # Importación es el más barato del catálogo actual
    assert min(comparacion, key=comparacion.get) == "Importacion"


def test_costeo_sustituye_sku_por_proveedor(proyecto):
    c = costear(proyecto, Tarifas(), proveedor="Arauco")
    assert c["sustituciones"].get("MEL-BLA-15-IMP") == "MEL-BLA-15-ARA"
